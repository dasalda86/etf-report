from pykrx import stock
import pandas as pd
from datetime import datetime, timedelta
import os
from openpyxl.styles import PatternFill

# 1. 저장 경로 설정 (깃허브 저장소 내 data 폴더)
save_path = "data"
if not os.path.exists(save_path): 
    os.makedirs(save_path)

# 2. 날짜 설정 (오늘 기준)
now = datetime.now()
# 깃허브 서버는 UTC 기준이므로 한국 시간(UTC+9)으로 보정 로직 (선택사항이나 권장)
# 한국 시간으로 밤 8시 30분 실행 시, 날짜가 꼬이지 않도록 보정
now_korea = now + timedelta(hours=9)
today_str = now_korea.strftime("%Y%m%d")
weekday_list = ['월', '화', '수', '목', '금', '토', '일']
today_weekday = weekday_list[now_korea.weekday()]

# 지난주 금요일 계산 (비교용)
days_to_last_friday = (now_korea.weekday() - 4) % 7
if days_to_last_friday == 0: days_to_last_friday = 7
last_friday_date = (now_korea - timedelta(days=days_to_last_friday)).strftime("%Y%m%d")

# 3. 섹터 및 ETF 맵핑 (기존과 동일)
sector_map = {
    "반도체": {"388420": "RISE 비메모리반도체액티브", "474590": "WON 반도체밸류체인액티브", "494220": "UNICORN SK하이닉스밸류체인액티브"},
    "수급 및 배당성장": {"0088N0": "WON K-글로벌수급상위", "444200": "SOL 코리아메가테크액티브", "476850": "KoAct 배당성장액티브", "441800": "TIME Korea플러스배당액티브"},
    "신재생 및 2차전지": {"385510": "KODEX 신재생에너지액티브", "404120": "TIME K신재생에너지액티브", "422420": "RISE 2차전지액티브"},
    "AI인프라 및 소비": {"385710": "TIME K이노베이션액티브", "422260": "VITA MZ소비액티브", "487130": "KoAct AI인프라액티브"},
    "코스피, 조선, 테크": {"385720": "TIME 코스피액티브", "445150": "KODEX 친환경조선해운액티브", "471780": "TIGER 코리아테크액티브"},
    "수출, 로봇, 컬쳐": {"0074K0": "KoAct K수출핵심기업TOP30액티브", "445290": "KODEX 로봇액티브", "410870": "TIME K컬처액티브"},
    "밸류업": {"495060": "TIME 코리아밸류업액티브", "495230": "KoAct 코리아밸류업액티브", "496130": "TRUSTON 코리아밸류업액티브"},
    "바이오": {"463050": "TIME K바이오액티브", "462900": "KoAct 바이오헬스케어액티브", "0000Z0": "RISE 바이오TOP10액티브"}
}

# 스타일 정의
orange_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
green_fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')

sheet1_list, sheet2_list, sheet3_list = [], [], []

# 4. 분석 엔진
for sector, etfs in sector_map.items():
    for ticker, etf_name in etfs.items():
        try:
            df_today = stock.get_etf_portfolio_deposit_file(ticker, today_str)
            df_l_fri = stock.get_etf_portfolio_deposit_file(ticker, last_friday_date)
            
            if df_today.empty: continue
            
            n_col = next((c for c in ['계정명', '종목명', '구성종목명'] if c in df_today.columns), None)
            w_col = next((c for c in ['비중', '구성비중'] if c in df_today.columns), None)
            
            # 오늘 데이터 저장 (요일별 csv)
            df_today[n_col] = df_today[n_col].astype(str).str.strip()
            df_today[[n_col, w_col]].to_csv(f"{save_path}/{ticker}_{today_weekday}.csv", index=False)

            # [시트 1] 요일별 흐름 통합
            df_l_fri_sub = df_l_fri[[n_col, w_col]].copy() if not df_l_fri.empty else pd.DataFrame(columns=[n_col, w_col])
            df_l_fri_sub.columns = [n_col, '전주금(%)']
            df_l_fri_sub['전주금(%)'] = df_l_fri_sub['전주금(%)'].astype(float).round(2)
            
            flow = df_l_fri_sub.copy()
            daily_dfs = {}
            for wd in ['월', '화', '수', '목', '금']:
                path = f"{save_path}/{ticker}_{wd}.csv"
                if os.path.exists(path):
                    temp = pd.read_csv(path)
                    temp.columns = [n_col, f'{wd}(%)']
                    temp[f'{wd}(%)'] = temp[f'{wd}(%)'].astype(float).round(2)
                    flow = pd.merge(flow, temp, on=n_col, how='outer')
                    daily_dfs[wd] = temp.set_index(n_col)
            
            flow['섹터'], flow['ETF명'] = sector, etf_name
            sheet1_list.append(flow)

            # [시트 2 & 3] 주간 변동 및 핵심 매수 분석
            df_now_idx = df_today[[n_col, w_col]].set_index(n_col)
            df_old_idx = df_l_fri_sub.set_index(n_col)
            
            for s in (set(df_now_idx.index) | set(df_old_idx.index)):
                v_old = round(float(df_old_idx.loc[s, '전주금(%)']), 2) if s in df_old_idx.index else 0.0
                v_now = round(float(df_now_idx.loc[s, w_col]), 2) if s in df_now_idx.index else 0.0
                diff = round(v_now - v_old, 2)
                
                status = "신규" if v_old == 0 else ("편출" if v_now == 0 else ("증가" if diff > 0 else "감소"))

                if abs(diff) >= 0.01:
                    sheet2_list.append({'섹터': sector, 'ETF명': etf_name, '종목명': s, '지난주금요일(%)': v_old, f'오늘({today_weekday})(%)': v_now, '주간변동(%p)': diff, '상태': status})

                if status == "신규" or diff >= 0.5:
                    # 액션 요일 찾기
                    action_day = "점진적"
                    for wd in ['월', '화', '수', '목', '금']:
                        if wd in daily_dfs and s in daily_dfs[wd].index:
                            if status == "신규": action_day = f"{wd} 신규"; break
                            else: action_day = f"{wd} 집중매수" # 단순화
                    
                    sheet3_list.append({'섹터': sector, '종목명': s, 'ETF명': etf_name, '증가한 비중(%p)': diff, '액션요일': action_day})
        except: continue

# 5. 엑셀 저장
report_filename = f"{save_path}/주간_ETF_분석리포트_{today_str}.xlsx"
with pd.ExcelWriter(report_filename, engine='openpyxl') as writer:
    if sheet1_list:
        pd.concat(sheet1_list).fillna(0).to_excel(writer, sheet_name='요일별_비중흐름', index=False)
    if sheet2_list:
        pd.DataFrame(sheet2_list).to_excel(writer, sheet_name='주간_변동분석', index=False)
    if sheet3_list:
        pd.DataFrame(sheet3_list).sort_values(['섹터', '증가한 비중(%p)'], ascending=[True, False]).to_excel(writer, sheet_name='핵심_매수_종목', index=False)
    
    # 스타일 적용 (색상)
    for sn in writer.sheets:
        ws = writer.sheets[sn]
        if sn == '요일별_비중흐름':
            for r in range(2, ws.max_row+1):
                for c in range(5, ws.max_column+1):
                    if ws.cell(r, c).value and ws.cell(r, c-1).value:
                        if ws.cell(r, c).value > ws.cell(r, c-1).value: ws.cell(r, c).fill = orange_fill
                        elif ws.cell(r, c).value < ws.cell(r, c-1).value: ws.cell(r, c).fill = green_fill
        elif sn == '주간_변동분석':
            for r in range(2, ws.max_row+1):
                if ws.cell(r, 6).value:
                    if ws.cell(r, 6).value > 0: ws.cell(r, 6).fill = orange_fill
                    else: ws.cell(r, 6).fill = green_fill
        elif sn == '핵심_매수_종목':
            for r in range(2, ws.max_row+1):
                ws.cell(r, 4).fill = orange_fill
                ws.cell(r, 5).fill = orange_fill

print(f"✅ 리포트 생성 완료: {report_filename}")
